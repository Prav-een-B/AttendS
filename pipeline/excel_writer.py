"""
pipeline/excel_writer.py
-------------------------
Generates a formatted .xlsx file from parsed attendance data.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


# ──────────────────────────────────────────────────────────────────────────────
# Colour palette (warm parchment + dark ink aesthetic)
# ──────────────────────────────────────────────────────────────────────────────

C_HEADER_BG  = "1C3557"   # dark blue
C_HEADER_FG  = "FFFFFF"
C_META_BG    = "EEF2FF"
C_META_FG    = "1C3557"
C_PRESENT_BG = "DCFCE7"   # mint green
C_ABSENT_BG  = "FEE2E2"   # light red
C_LATE_BG    = "FEF9C3"   # light yellow
C_ALT_BG     = "F8FAFF"   # very light blue
C_WHITE      = "FFFFFF"
C_ACCENT     = "2563EB"
C_MUTED      = "64748B"


def _thin_border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def _font(bold=False, size=10, color="000000", name="Calibri"):
    return Font(name=name, bold=bold, size=size, color=color)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ──────────────────────────────────────────────────────────────────────────────
# Sheet builder
# ──────────────────────────────────────────────────────────────────────────────

def _write_sheet(ws, parsed: dict, source_label: str):
    meta    = parsed.get("meta", {})
    records = parsed.get("records", [])
    stats   = parsed.get("stats", {})
    conf    = parsed.get("confidence", "Medium")

    border = _thin_border()

    row = 1

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:G{row}")
    c = ws[f"A{row}"]
    c.value     = meta.get("title") or "Attendance Sheet"
    c.font      = _font(bold=True, size=14, color=C_HEADER_FG)
    c.fill      = _fill(C_HEADER_BG)
    c.alignment = _align()
    ws.row_dimensions[row].height = 32
    row += 1

    # ── Meta row ───────────────────────────────────────────────────────────────
    meta_items = [
        ("Date",        meta.get("date")    or "—"),
        ("Subject",     meta.get("subject") or "—"),
        ("Source",      source_label),
        ("Confidence",  conf),
    ]
    col_spans = [(1, 2), (3, 4), (5, 6), (7, 7)]
    for (label, value), (c1, c2) in zip(meta_items, col_spans):
        if c1 != c2:
            ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row=row, column=c1)
        cell.value     = f"{label}: {value}"
        cell.font      = _font(bold=True, size=9, color=C_META_FG)
        cell.fill      = _fill(C_META_BG)
        cell.alignment = _align(h="left")
        cell.border    = border
    ws.row_dimensions[row].height = 18
    row += 1

    # spacer
    row += 1

    # ── Column headers ─────────────────────────────────────────────────────────
    headers = ["#", "Roll Number", "Student Name", "Status", "Serial", "Remarks", "Raw OCR"]
    col_widths = [5, 16, 32, 12, 8, 18, 40]
    for c_idx, (h, cw) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=row, column=c_idx, value=h)
        cell.font      = _font(bold=True, size=10, color=C_HEADER_FG)
        cell.fill      = _fill(C_HEADER_BG)
        cell.alignment = _align()
        cell.border    = border
        ws.column_dimensions[get_column_letter(c_idx)].width = cw
    ws.row_dimensions[row].height = 20
    row += 1

    # ── Data rows ──────────────────────────────────────────────────────────────
    for r_idx, rec in enumerate(records, start=1):
        status = rec.get("status", "Unknown")
        if   status == "Present": bg = C_PRESENT_BG
        elif status == "Absent":  bg = C_ABSENT_BG
        elif status == "Late":    bg = C_LATE_BG
        else:                     bg = C_ALT_BG if r_idx % 2 == 0 else C_WHITE

        status_color = {
            "Present": "166534", "Absent": "991B1B",
            "Late": "92400E",    "Unknown": C_MUTED,
        }.get(status, C_MUTED)

        values = [
            r_idx,
            rec.get("roll_number", ""),
            rec.get("name", ""),
            status,
            rec.get("serial", ""),
            "",                          # Remarks — left blank for manual fill
            rec.get("raw", "")[:80],
        ]
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c_idx, value=val)
            cell.fill   = _fill(bg)
            cell.border = border
            cell.alignment = _align(h="left" if c_idx in (3, 7) else "center")
            if c_idx == 4:  # Status column
                cell.font = _font(bold=True, size=10, color=status_color)
            else:
                cell.font = _font(size=10)
        ws.row_dimensions[row].height = 16
        row += 1

    # ── Summary ────────────────────────────────────────────────────────────────
    row += 1
    summary_items = [
        ("Total Students",  stats.get("total",   0)),
        ("Present",         stats.get("present", 0)),
        ("Absent",          stats.get("absent",  0)),
        ("Late",            stats.get("late",    0)),
        ("Unknown",         stats.get("unknown", 0)),
        ("Generated",       datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for label, value in summary_items:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=4, value=str(value))
        lc.font = _font(bold=True, size=9, color=C_ACCENT)
        vc.font = _font(size=9)
        lc.alignment = vc.alignment = _align(h="left")
        row += 1


# ──────────────────────────────────────────────────────────────────────────────
# Public API
# ──────────────────────────────────────────────────────────────────────────────

def write_excel(all_parsed: list, output_path: str):
    """
    all_parsed: list of dicts, each from parser.parse_ocr_output(),
                with an extra '_source_file' key added by app.py.
    """
    wb = openpyxl.Workbook()
    first = True

    for idx, parsed in enumerate(all_parsed):
        source = parsed.get("_source_file", f"Image {idx+1}")
        sheet_title = (parsed.get("meta", {}).get("title") or f"Sheet {idx+1}")[:31]

        if first:
            ws = wb.active
            ws.title = sheet_title
            first = False
        else:
            ws = wb.create_sheet(title=sheet_title)

        _write_sheet(ws, parsed, source)

    # ── Multi-image summary sheet ──────────────────────────────────────────────
    if len(all_parsed) > 1:
        ws_s = wb.create_sheet(title="Summary")
        border = _thin_border()

        ws_s.merge_cells("A1:F1")
        c = ws_s["A1"]
        c.value = "Attendance Summary — All Sheets"
        c.font  = _font(bold=True, size=13, color=C_HEADER_FG)
        c.fill  = _fill(C_HEADER_BG)
        c.alignment = _align()
        ws_s.row_dimensions[1].height = 28

        hdrs = ["#", "Sheet / Title", "Total", "Present", "Absent", "Confidence"]
        widths = [5, 35, 10, 10, 10, 14]
        for c_idx, (h, w) in enumerate(zip(hdrs, widths), start=1):
            cell = ws_s.cell(row=2, column=c_idx, value=h)
            cell.font = _font(bold=True, size=10, color=C_HEADER_FG)
            cell.fill = _fill(C_HEADER_BG)
            cell.alignment = _align()
            cell.border = border
            ws_s.column_dimensions[get_column_letter(c_idx)].width = w

        for i, parsed in enumerate(all_parsed, start=1):
            stats = parsed.get("stats", {})
            row_vals = [
                i,
                parsed.get("meta", {}).get("title") or f"Sheet {i}",
                stats.get("total", 0),
                stats.get("present", 0),
                stats.get("absent", 0),
                parsed.get("confidence", "?"),
            ]
            bg = C_ALT_BG if i % 2 == 0 else C_WHITE
            for c_idx, v in enumerate(row_vals, start=1):
                cell = ws_s.cell(row=i+2, column=c_idx, value=v)
                cell.font = _font(size=10)
                cell.fill = _fill(bg)
                cell.alignment = _align()
                cell.border = border

    wb.save(output_path)
